#!/usr/bin/env python3
"""Genera la plantilla buida del quadern digital del docent (.xlsx).

Crea `Avaluació/Quadern_digital_docent_plantilla.xlsx` amb les pestanyes, validacions i
formats descrits a `Avaluació/Quadern_digital_docent.md`. La plantilla NO conté dades
personals: es puja al Drive de centre i s'hi posen els noms ALLÀ (mai al repositori).

Ús:  python genera_quadern.py     Requereix: pip install openpyxl
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).parent
DESTI = ROOT / "Avaluació" / "Quadern_digital_docent_plantilla.xlsx"

N_ALUMNES = 20
TARONJA = "E8641B"
CREMA = "FDEADD"

CES = ["CE1", "CE2", "CE3", "CE4", "CE5", "CE6", "TE", "TA"]
SAS = [f"SA{i}" for i in range(1, 10)]
NIVELLS = '"NA,AS,AN,AE"'


def capcalera(ws, cols, amplades=None):
    for i, nom in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=nom)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=TARONJA)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if amplades:
            ws.column_dimensions[get_column_letter(i)].width = amplades[i - 1]
    ws.freeze_panes = "A2"


wb = Workbook()

# ── 1. Alumnat ─────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Alumnat"
capcalera(ws, ["Alumne/a", "Grup (A/B)", "Equip T1", "Equip T2", "Equip T3",
               "Perfil / NESE / adaptacions"],
          [26, 10, 12, 12, 12, 45])
dv = DataValidation(type="list", formula1='"A,B"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"B2:B{N_ALUMNES + 1}")

# ── 2. Seguiment ───────────────────────────────────────────────────────────
ws = wb.create_sheet("Seguiment")
capcalera(ws, ["Data", "Alumne/a", "Docent", "Codi", "CA", "Nota breu"],
          [12, 26, 10, 7, 8, 60])
dv = DataValidation(type="list", formula1='"Maker,Co"', allow_blank=True)
ws.add_data_validation(dv); dv.add("C2:C500")
dv = DataValidation(type="list", formula1='"+,-,!"', allow_blank=True)
ws.add_data_validation(dv); dv.add("D2:D500")
ws.cell(row=2, column=1).comment = None

# ── 3. Progrés ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Progrés")
cols = ["Alumne/a"] + [f"{sa}·{ce}" for sa in SAS for ce in CES] + \
       [f"Global {ce}" for ce in CES] + ["T1", "T2", "T3"]
capcalera(ws, cols, [26] + [7] * (len(cols) - 1))
dv = DataValidation(type="list", formula1=NIVELLS, allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"B2:{get_column_letter(len(cols))}{N_ALUMNES + 1}")
# format condicional: la trajectòria es veu de vermell a verd
colors = {"NA": "F4B8A0", "AS": "FCE8B2", "AN": "D9EAD3", "AE": "93C47D"}
rang = f"B2:{get_column_letter(len(cols))}{N_ALUMNES + 1}"
for nivell, color in colors.items():
    ws.conditional_formatting.add(
        rang, CellIsRule(operator="equal", formula=[f'"{nivell}"'],
                         fill=PatternFill("solid", fgColor=color)))

# ── 4. Carnets ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Carnets")
capcalera(ws, ["Alumne/a",
               "Làser (data)", "3D (data)", "360 (data)", "VR (data)",
               "Usos làser T1", "Usos làser T2", "Usos làser T3",
               "Usos 3D T1", "Usos 3D T2", "Usos 3D T3",
               "Usos 360 T3", "Usos VR T3", "Observacions"],
          [26] + [11] * 4 + [9] * 8 + [30])

# ── 5. Equitat (fórmules d'exemple sobre Carnets) ──────────────────────────
ws = wb.create_sheet("Equitat")
ws["A1"] = "Revisió trimestral d'equitat — es calcula sola des de la pestanya Carnets"
ws["A1"].font = Font(bold=True, size=12)
ws["A3"] = ("Com fer-la: inseriu una taula dinàmica sobre Carnets (files: alumne/a o grup; "
            "valors: suma d'usos per màquina i trimestre), o useu les fórmules d'exemple:")
ws["A5"] = "Total d'usos de làser T1 del grup:"
ws["C5"] = "=SUM(Carnets!F2:F21)"
ws["A6"] = "Alumnes amb ZERO usos de cap màquina al T1:"
ws["C6"] = '=COUNTIFS(Carnets!F2:F21,0,Carnets!I2:I21,0)'
ws["A8"] = ("Pregunta clau: hi ha alumnes que no han operat mai res? El repartiment reflecteix "
            "el grup o el biaix «ell talla / ella documenta»? → correcció via criteri d'ordre "
            "del kanban (vegeu Equitat_genere_STEAM.md §3·bis).")
for f in ("A1", "A3", "A8"):
    ws[f].alignment = Alignment(wrap_text=True)
ws.column_dimensions["A"].width = 60
ws.column_dimensions["C"].width = 30

# ── 6. Kanban ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("Kanban")
capcalera(ws, ["Peça / placa (nom de fitxer!)", "Alumnes", "Estat", "Temps est.",
               "Grams", "Data enviat", "Data fet"],
          [34, 24, 14, 10, 8, 12, 12])
dv = DataValidation(type="list", formula1='"Per imprimir,Imprimint,Fet,Fallida"',
                    allow_blank=True)
ws.add_data_validation(dv); dv.add("C2:C300")

# ── 7. Valoracions ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Valoracions")
ws["A1"] = "Veu de l'alumnat — tiquet trimestral anònim (Instruments_formatius.md §8)"
ws["A1"].font = Font(bold=True, size=12)
capcalera_fila3 = ["Trimestre", "Mantindria…", "Trauria/canviaria…", "M'agradaria fer…",
                   "Conclusió i retorn donat al grup"]
for i, nom in enumerate(capcalera_fila3, 1):
    c = ws.cell(row=3, column=i, value=nom)
    c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor=CREMA)
ws.column_dimensions["A"].width = 10
for col in "BCDE":
    ws.column_dimensions[col].width = 40

# ── full de portada amb l'avís RGPD ────────────────────────────────────────
ws = wb.create_sheet("LLEGIU-ME", 0)
ws["A1"] = "QUADERN DIGITAL DEL DOCENT — Aula Maker 1r ESO"
ws["A1"].font = Font(bold=True, size=14, color=TARONJA)
ws["A3"] = ("⚠️ RGPD: aquest fitxer, un cop hi poseu NOMS, conté dades personals d'alumnes. "
            "Pugeu-lo al Drive DE CENTRE i compartiu-lo només amb l'altre docent. "
            "MAI el pugeu al repositori públic (allà només hi viu la plantilla buida).")
ws["A5"] = ("Ús: ompliu la pestanya Alumnat; Seguiment és el registre diari (+/-/!); "
            "Progrés s'omple en tancar cada SA (colors automàtics NA→AE); Carnets registra "
            "carnets i usos; Equitat es llegeix al tancament de trimestre; Kanban és el "
            "registre de fabricació; Valoracions recull la veu de l'alumnat.")
ws["A7"] = "Especificació completa: Avaluació/Quadern_digital_docent.md"
for f in ("A3", "A5"):
    ws[f].alignment = Alignment(wrap_text=True)
ws.column_dimensions["A"].width = 100

wb.save(DESTI)
print(f"Plantilla generada: {DESTI}")
